import pandas as pd
import numpy as np
from sklearn.linear_model import LinearRegression
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


file_path = "corrected_merged_dataset.xlsx"
df = pd.read_excel(file_path)


ct_c0_cols_TBA = ["c6/c0", "c7/c0", "c8/c0", "c9/c0", "c10/c0"]
t_cols_TBA = ["t6", "t7", "t8", "t9", "t10"]
ct_c0_cols_MeOH = ["c11/c0", "c12/c0", "c13/c0", "c14/c0", "c15/c0"]
t_cols_MeOH = ["t11", "t12", "t13", "t14", "t15"]


for col in ["kTBA", "kMeOH", "kkk", "kMeOH/kkk", "kTBA/kkk", "1-kMeOH/kkk", "1-kTBA/kkk"]:
    if col not in df.columns:
        df[col] = np.nan


def linear_fit(ct_vals, t_vals):
    x = t_vals.reshape(-1, 1)
    y = -np.log(ct_vals)
    model = LinearRegression().fit(x, y)
    return model.coef_[0], model.score(x, y)


rows_to_highlight = set()

for i in range(len(df)):
    try:
        if pd.notna(df.at[i, "c6/c0"]):
            ct_vals = df.loc[i, ct_c0_cols_TBA].astype(float).values
            t_vals = df.loc[i, t_cols_TBA].astype(float).values
            k, r2 = linear_fit(ct_vals, t_vals)
            df.at[i, "kTBA"] = k
            if r2 < 0.85:
                rows_to_highlight.add(i)

        if pd.notna(df.at[i, "c11/c0"]):
            ct_vals = df.loc[i, ct_c0_cols_MeOH].astype(float).values
            t_vals = df.loc[i, t_cols_MeOH].astype(float).values
            k, r2 = linear_fit(ct_vals, t_vals)
            df.at[i, "kMeOH"] = k
            if r2 < 0.85:
                rows_to_highlight.add(i)
    except:
        continue


for i in range(len(df)):
    if df.loc[i, ["kTBA", "kMeOH", "c6/c0", "c11/c0"]].notna().any():
        ref_row = df.iloc[i, 0:58]
        for j in range(len(df)):
            if i != j and df.iloc[j, 0:58].equals(ref_row):
                if pd.notna(df.at[j, "k"]):
                    df.at[i, "kkk"] = df.at[j, "k"]
                break


df["kMeOH/kkk"] = df["kMeOH"] / df["kkk"]
df["kTBA/kkk"] = df["kTBA"] / df["kkk"]
df["kTBA/kkk-kMeOH/kkk"] = df["kTBA/kkk"] - df["kMeOH/kkk"]
df["1-kTBA/kkk"] = 1 - df["kTBA/kkk"]


output_file = "corrected_with_kTBA_kMeOH_highlight.xlsx"
df.to_excel(output_file, index=False)


wb = load_workbook(output_file)
ws = wb.active
red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")

for row_idx in rows_to_highlight:
    for cell in ws[row_idx + 2]:
        cell.fill = red_fill

wb.save(output_file)
print(f"\nsaved as {output_file}， R² <0.85 red。")
