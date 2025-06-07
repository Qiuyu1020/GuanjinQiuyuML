import pandas as pd
import numpy as np
from sklearn.linear_model import LinearRegression
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


df = pd.read_excel("dataset_filled_elements_ions_0.xlsx")


# ct_c0_cols = ["c1/c0", "c2/c0", "c3/c0", "c4/c0", "c5/c0"]
# t_cols = ["t1", "t2", "t3", "t4", "t5"]

ct_c0_cols = ["c6/c0", "c7/c0", "c8/c0", "c9/c0", "c10/c0"]
t_cols = ["t6", "t7", "t8", "t9", "t10"]

highlight_indices = []
omega = 0

for i in range(len(df)):
    try:
        if pd.notna(df.at[i, "c6/c0"]):
            ct_vals = df.loc[i, ct_c0_cols].astype(float).values
            t_vals = df.loc[i, t_cols].astype(float).values
            x = t_vals.reshape(-1, 1)
            y = -np.log(ct_vals)

            model = LinearRegression().fit(x, y)
            k = model.coef_[0]
            r2 = model.score(x, y)

            df.at[i, "kTBA"] = k

            if r2 < 0.85:

                highlight_indices.append(i)
                omega = omega + 1
                print(f"R^2 = {r2:.4f}" + f"numbers of highlight data: = {omega: .4f}")


    except Exception:
        continue

def compute_log2k(x):
    if pd.notna(x) and isinstance(x, (int, float)) and x > 0:
        return np.log2(x)
    else:
        return ""


df["kTBA/k"] = df["k"].apply(compute_log2k)


output_file = "dataset_filled_elements_ions_0_k_log2k_ktba.xlsx"
df.to_excel(output_file, index=False)


wb = load_workbook(output_file)
ws = wb.active
red_fill = PatternFill(start_color="FFFF9999", end_color="FFFF9999", fill_type="solid")

for i in highlight_indices:
    for cell in ws[i + 2]:
        cell.fill = red_fill

wb.save(output_file)

print(f"saved as：{output_file}")