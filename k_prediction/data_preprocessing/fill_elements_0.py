import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# read data
df = pd.read_excel("dataset1.xlsx")

# all 118 elements
basic_elements = {
    'H', 'He', 'Li', 'Be', 'B', 'C', 'N', 'O', 'F', 'Ne',
    'Na', 'Mg', 'Al', 'Si', 'P', 'S', 'Cl', 'Ar',
    'K', 'Ca', 'Sc', 'Ti', 'V', 'Cr', 'Mn', 'Fe', 'Co', 'Ni', 'Cu', 'Zn',
    'Ga', 'Ge', 'As', 'Se', 'Br', 'Kr',
    'Rb', 'Sr', 'Y', 'Zr', 'Nb', 'Mo', 'Tc', 'Ru', 'Rh', 'Pd', 'Ag', 'Cd',
    'In', 'Sn', 'Sb', 'Te', 'I', 'Xe',
    'Cs', 'Ba', 'La', 'Ce', 'Pr', 'Nd', 'Pm', 'Sm', 'Eu', 'Gd', 'Tb', 'Dy',
    'Ho', 'Er', 'Tm', 'Yb', 'Lu',
    'Hf', 'Ta', 'W', 'Re', 'Os', 'Ir', 'Pt', 'Au', 'Hg',
    'Tl', 'Pb', 'Bi', 'Po', 'At', 'Rn',
    'Fr', 'Ra', 'Ac', 'Th', 'Pa', 'U', 'Np', 'Pu', 'Am', 'Cm', 'Bk', 'Cf',
    'Es', 'Fm', 'Md', 'No', 'Lr',
    'Rf', 'Db', 'Sg', 'Bh', 'Hs', 'Mt', 'Ds', 'Rg', 'Cn', 'Fl', 'Lv', 'Ts', 'Og'
}

# copy dataset
df_processed = df.copy()

# normalize element columns
for col in df.columns:
    if col in basic_elements:
        df_processed[col] = df[col].fillna(0)
        df_processed[col] = df_processed[col].apply(lambda x: x / 100 if x > 1 else x)

# save temp file
output_path = "dataset_filled_elements_0.xlsx"
df_processed.to_excel(output_path, index=False)

# reload with openpyxl
wb = load_workbook(output_path)
ws = wb.active

# style for red background
red_fill = PatternFill(start_color="FFFF9999", end_color="FFFF9999", fill_type="solid")

# get column indices for elements
element_cols = [i + 1 for i, col in enumerate(df_processed.columns) if col in basic_elements]

# iterate and check row sums
print("sum of elements row not equal to 1：")
for row in range(2, ws.max_row + 1):
    row_sum = sum(ws.cell(row=row, column=col).value or 0 for col in element_cols)
    if abs(row_sum - 1) > 1e-2:
        print(f"row {row}: elements mass ratio sum = {row_sum:.6f}")
        for col in range(1, ws.max_column + 1):
            ws.cell(row=row, column=col).fill = red_fill

# save updated file
wb.save(output_path)
print("saved as：", output_path)
