# 批量处理得到数据集中的k
# 若R^2小于0.9则对应的k空着
import pandas as pd
import numpy as np
from sklearn.linear_model import LinearRegression
from sklearn.metrics import r2_score
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


file_path = 'D:\\DESKTOP\\Guanjin_ML\\data_preparation\\d1.xlsx'
df = pd.read_excel(file_path)


k_values = []


for idx, row in df.iterrows():
    try:

        C0 = row['C0']
        Ct_list = [row['Ct1'], row['Ct2'], row['Ct3'], row['Ct4'], row['Ct5']]
        t_list = [row['t1'], row['t2'], row['t3'], row['t4'], row['t5']]

        Ct_array = np.array(Ct_list)
        t_array = np.array(t_list)

        # y = -ln(Ct/C0)
        y = -np.log(Ct_array / C0)
        x = t_array.reshape(-1, 1)


        model = LinearRegression()
        model.fit(x, y)
        k = model.coef_[0]
        y_pred = model.predict(x)
        r2 = r2_score(y, y_pred)

       ###################### # R^2 boundary condition
        if r2 >= 0.9:
            k_values.append(k)
        else:
            k_values.append(np.nan)
    except Exception as e:
        print(f"第{idx+2}行出错: {e}")
        k_values.append(np.nan)


df['k'] = k_values

# 保存初步Excel
save_path = 'D:\DESKTOP\Guanjin_ML\Guanjin-Qiuyu-Machine-Learning\d1_fitted_highlight.xlsx'
df.to_excel(save_path, index=False)

wb = load_workbook(save_path)
ws = wb.active


red_fill = PatternFill(start_color='FF9999', end_color='FF9999', fill_type='solid')


for row_idx in range(2, ws.max_row + 1):
    k_cell = ws[f'Q{row_idx}']
    if k_cell.value is None or (isinstance(k_cell.value, float) and np.isnan(k_cell.value)):
        for col_idx in range(1, ws.max_column + 1):
            ws.cell(row=row_idx, column=col_idx).fill = red_fill


wb.save(save_path)

print(f"批量拟合并高亮完成，已保存到：{save_path}")
